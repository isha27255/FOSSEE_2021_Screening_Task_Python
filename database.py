import sqlite3
import xlrd as xl
# python library to read/write excel files
import openpyxl

dbname = "steel_sections.sqlite"

def start_connection():
    """Establish connection with Database"""
    global dbname
    conn = sqlite3.connect(dbname)
    return conn

def connect(func):
    def inner(conn, *args, **kwargs):
        query = 'SELECT name FROM sqlite_master WHERE type="table";'
        try:
            conn.execute(query)
        except Exception as e:
            conn = start_connection()
        return func(conn, *args, **kwargs)
    return inner

@connect
def insert_upon_delete(conn, table_name, values):
    """Updates the value by first deleting the previous value and then inserting new."""
    if table_name == 'Beams':
        #20
        query = 'INSERT INTO Beams VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    elif table_name == 'Angles':
        #24
        query = 'INSERT INTO Angles VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    else:
        #21
        query = 'INSERT INTO Channels VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    conn.execute(query, values)
    conn.commit()

@connect
def insert_one(conn, table_name, values):
    """Inserts data into table mentioned in the `table_name`"""
    if table_name == 'Beams':
        #20
        query = 'INSERT INTO Beams VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    elif table_name == 'Angles':
        #24
        query = 'INSERT INTO Angles VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    else:
        #21
        query = 'INSERT INTO Channels VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'

    query2 = "SELECT designation FROM {} WHERE designation=?".format(table_name)
    present = conn.execute(query2, (values[0],)).fetchone()
    if present:
        raise ItemAlreadyStored()
    id = get_Id(conn, table_name)
    values.insert(0, int(id)+1)
    conn.execute(query, values)
    conn.commit()



@connect
def insert_many(conn, table_name, values):
    if table_name == 'Beams':
        #20
        query = 'INSERT INTO Beams VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    elif table_name == 'Angles':
        #24
        query = 'INSERT INTO Angles VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    else:
        #21
        query = 'INSERT INTO Channels VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    for value in values:
        conn.execute(query, value)
    conn.commit()

@connect
def view_all(conn, table_name):
    """View Table data of `table_name`"""
    if table_name is None:
        raise NotValidTable("Enter a valid table name")
    query = "SELECT * FROM {}".format(table_name)
    return conn.execute(query)

@connect
def view_one(conn, table_name, designation):
    if table_name is None:
        raise NotValidTable("Enter a valid table name")
    query = "SELECT * FROM {} WHERE designation=?".format(table_name)
    return conn.execute(query, (designation,))

@connect
def update_one(conn, table_name, values, designation):
    query = "SELECT Id FROM {} WHERE designation=?".format(table_name)
    result = conn.execute(query, (designation,)).fetchone()
    query2 = "DELETE FROM {} WHERE Id=?".format(table_name)
    conn.execute(query2, result)
    values.insert(0, result[0])
    insert_upon_delete(conn, table_name, values)

@connect
def delete_one(conn, table_name, designation):
    query = "DELETE FROM {} WHERE designation=?".format(table_name)
    conn.execute(query, (designation,))
    conn.commit()

@connect
def get_designations(conn, table, *args, **kwargs):
    result = list()
    query = "SELECT Designation from {}".format(table)
    for i in conn.execute(query):
        result.append(i[0])
    return result

@connect
def get_columns(conn, table_name):
    column_names = list()
    for row in conn.execute('PRAGMA table_info="{}"'.format(table_name)):
        column_names.append(row[1])
    return column_names



def get_Id(conn, table, *args, **kwargs):
    query = "SELECT MAX(Id) from {}".format(table)
    result = conn.execute(query).fetchone()
    return result[0]

# @connect
# def insert_using_excel(conn, table, loc):
#     wb = xl.open_workbook(loc)
#     s1 = wb.sheet_by_index(0) 
#     print("No. of rows:", s1.nrows)
#     print("No. of columns:", s1.ncols)  


@connect
def insert_using_excel(conn, table, loc):
    wb = openpyxl.load_workbook(loc)
    if table == 'Angles':
        s = 1
    elif table == 'Channels':
        s=2
    else:
        s=0
    wb.active = s
    sheet = wb.active
    # print(sheet)
    row_count = sheet.max_row
    col_count = sheet.max_column
    # print(table, row_count, col_count)
    max_id = get_Id(conn, table)
    values = list()
    for x in range(2, row_count+1):
        temp = list()
        for y in range(2, col_count+1):
            cell_obj = sheet.cell(row=x, column=y)
            item_val = cell_obj.value
            
            try:
                item_val = float(item_val)
            except ValueError:
                pass
            temp.append(item_val)
        temp.insert(0, max_id+x-1)
        values.append(temp)
    insert_many(conn, table, values)

